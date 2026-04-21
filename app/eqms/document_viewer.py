"""
Centralized document viewer — renders file types that browsers can't display natively.

Handles:
  .docx  → HTML via mammoth
  .xlsx / .xls → HTML table via openpyxl
  .csv   → HTML table

All other file types (PDF, images, text, .eml) should be served directly via send_file.
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from flask import Response, current_app, flash, redirect, render_template, url_for


# ---------------------------------------------------------------------------
# Public helpers
# ---------------------------------------------------------------------------

def needs_server_render(filename: str) -> bool:
    """Return True if the file type requires server-side rendering to be viewed in-browser.

    Note: ``.doc`` (legacy Word) is excluded — mammoth cannot convert it.
    Those files will be offered as downloads instead.
    """
    ext = Path(filename or "").suffix.lower()
    return ext in {".docx", ".xlsx", ".xls", ".csv"}


def render_document_to_response(
    file_bytes: bytes,
    filename: str,
    content_type: str,
    download_url: str,
    back_url: str | None = None,
    *,
    pdf_url: str | None = None,
    table_fallback_url: str | None = None,
    viewer_template: str = "admin/document_viewer.html",
) -> Response | None:
    """
    Convert *file_bytes* to an HTML page the browser can display.

    Returns a Flask ``Response`` on success, or ``None`` if the file type
    is not handled (caller should fall back to ``send_file``).
    """
    ext = Path(filename or "").suffix.lower()

    if ext == ".docx":
        return _render_docx(
            file_bytes,
            filename,
            download_url,
            back_url,
            pdf_url=pdf_url,
            table_fallback_url=table_fallback_url,
            viewer_template=viewer_template,
        )
    if ext in {".xlsx", ".xls"}:
        return _render_excel(
            file_bytes,
            filename,
            download_url,
            back_url,
            pdf_url=pdf_url,
            table_fallback_url=table_fallback_url,
            viewer_template=viewer_template,
        )
    if ext == ".csv":
        return _render_csv(file_bytes, filename, download_url, back_url)
    # Unsupported type — caller should fall back to download
    return None


# ---------------------------------------------------------------------------
# HTML sanitization (F-009)
# ---------------------------------------------------------------------------

# Tags considered safe for mammoth-converted content
_SAFE_TAGS = frozenset({
    "h1", "h2", "h3", "h4", "h5", "h6", "p", "br", "hr",
    "ul", "ol", "li", "dl", "dt", "dd",
    "table", "thead", "tbody", "tfoot", "tr", "th", "td", "caption", "colgroup", "col",
    "strong", "b", "em", "i", "u", "s", "sub", "sup", "small", "mark",
    "a", "img", "blockquote", "pre", "code", "span", "div",
})

_EVENT_ATTR_RE = re.compile(r"\s+on\w+\s*=", re.IGNORECASE)
_DANGEROUS_TAG_RE = re.compile(
    r"<\s*/?\s*(script|iframe|object|embed|form|input|button|textarea|select|meta|link|base|applet)\b[^>]*>",
    re.IGNORECASE,
)


def sanitize_viewer_html(html: str) -> str:
    """Public alias for mammoth / auditor PDF pipeline sanitization."""
    return _sanitize_html(html)


def _sanitize_html(html: str) -> str:
    """Strip dangerous tags and event-handler attributes from HTML.

    This is a defence-in-depth measure for mammoth output which is
    generally safe, but could contain malicious content from a crafted
    .docx file.
    """
    # Remove dangerous tags entirely
    html = _DANGEROUS_TAG_RE.sub("", html)
    # Remove on* event handler attributes (onclick, onerror, etc.)
    html = _EVENT_ATTR_RE.sub(" ", html)
    return html


# ---------------------------------------------------------------------------
# Private renderers
# ---------------------------------------------------------------------------

def _render_docx(
    file_bytes: bytes,
    filename: str,
    download_url: str,
    back_url: str | None,
    *,
    pdf_url: str | None = None,
    table_fallback_url: str | None = None,
    viewer_template: str = "admin/document_viewer.html",
) -> Response | None:
    try:
        import mammoth  # type: ignore
    except ImportError:
        current_app.logger.warning("mammoth not installed — cannot render .docx inline")
        return None

    try:
        result = mammoth.convert_to_html(io.BytesIO(file_bytes))
        html_body = result.value
        messages = result.messages
        if messages:
            current_app.logger.debug(
                "mammoth messages for %s: %s",
                filename,
                "; ".join(str(m) for m in messages),
            )
    except Exception as e:
        current_app.logger.error("Failed to convert %s to HTML: %s", filename, e)
        return None

    return _viewer_response(
        filename=filename,
        rendered_html=sanitize_viewer_html(html_body),
        download_url=download_url,
        back_url=back_url,
        pdf_url=pdf_url,
        table_fallback_url=table_fallback_url,
        viewer_template=viewer_template,
    )


def _render_excel(
    file_bytes: bytes,
    filename: str,
    download_url: str,
    back_url: str | None,
    *,
    pdf_url: str | None = None,
    table_fallback_url: str | None = None,
    viewer_template: str = "admin/document_viewer.html",
) -> Response | None:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        current_app.logger.warning("openpyxl not installed — cannot render Excel inline")
        return None

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheets: dict[str, list[list[str]]] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])
            if rows:
                sheets[sheet_name] = rows
        wb.close()
    except Exception as e:
        current_app.logger.error("Failed to parse Excel file %s: %s", filename, e)
        return None

    if not sheets:
        return None

    return _spreadsheet_response(
        filename=filename,
        sheets=sheets,
        download_url=download_url,
        back_url=back_url,
        pdf_url=pdf_url,
        table_fallback_url=table_fallback_url,
        viewer_template=viewer_template,
    )


def _render_csv(
    file_bytes: bytes,
    filename: str,
    download_url: str,
    back_url: str | None,
) -> Response | None:
    try:
        text = file_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = [row for row in reader]
    except Exception as e:
        current_app.logger.error("Failed to parse CSV %s: %s", filename, e)
        return None

    if not rows:
        return None

    sheets = {"Sheet1": rows}
    return _spreadsheet_response(
        filename=filename,
        sheets=sheets,
        download_url=download_url,
        back_url=back_url,
    )


# ---------------------------------------------------------------------------
# Response builders
# ---------------------------------------------------------------------------

def _viewer_response(
    filename: str,
    rendered_html: str,
    download_url: str,
    back_url: str | None,
    *,
    pdf_url: str | None = None,
    table_fallback_url: str | None = None,
    viewer_template: str = "admin/document_viewer.html",
) -> Response:
    """Wrap converted HTML content in the document viewer template."""
    return current_app.make_response(
        render_template(
            viewer_template,
            filename=filename,
            rendered_html=rendered_html,
            sheets=None,
            download_url=download_url,
            back_url=back_url,
            pdf_url=pdf_url,
            table_fallback_url=table_fallback_url,
        )
    )


def _spreadsheet_response(
    filename: str,
    sheets: dict[str, list[list[str]]],
    download_url: str,
    back_url: str | None,
    *,
    pdf_url: str | None = None,
    table_fallback_url: str | None = None,
    viewer_template: str = "admin/document_viewer.html",
) -> Response:
    """Wrap spreadsheet data in the document viewer template."""
    return current_app.make_response(
        render_template(
            viewer_template,
            filename=filename,
            rendered_html=None,
            sheets=sheets,
            download_url=download_url,
            back_url=back_url,
            pdf_url=pdf_url,
            table_fallback_url=table_fallback_url,
        )
    )
