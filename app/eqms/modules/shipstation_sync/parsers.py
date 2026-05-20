from __future__ import annotations

import csv
import os
import re
from pathlib import Path

from app.eqms.constants import EXCLUDED_SKUS, VALID_SKUS


def resolve_lotlog_path() -> str:
    """
    Resolve the absolute path to LotLog.csv.

    Priority:
    1. LOTLOG_PATH environment variable
    2. SHIPSTATION_LOTLOG_PATH environment variable
    3. LotLog_Path environment variable
    4. Default: app/eqms/data/LotLog.csv relative to project root
    """
    env_path = (
        os.environ.get("LOTLOG_PATH")
        or os.environ.get("SHIPSTATION_LOTLOG_PATH")
        or os.environ.get("LotLog_Path")
        or ""
    ).strip()
    if env_path:
        return env_path

    # Use absolute path based on project structure
    # parsers.py is at app/eqms/modules/shipstation_sync/parsers.py
    # project root is 4 levels up
    project_root = Path(__file__).resolve().parents[4]
    return str(project_root / "app" / "eqms" / "data" / "LotLog.csv")

# Regex patterns for lot extraction from text
LOT_RX = re.compile(r"\bSLQ-?\d+\b", re.IGNORECASE)
LOT_LABEL_RX = re.compile(r"LOT[:\s]*([A-Z0-9\-]+)", re.IGNORECASE)
# Bare numeric lot pattern (e.g., "05012025" in notes)
BARE_LOT_RX = re.compile(r"\b(\d{6,12})\b")
# Multi-SKU lot pattern: "SKU: 21600101003 LOT: SLQ-05012025"
SKU_LOT_PAIR_RX = re.compile(r"SKU[:\s]*(\d+)[^A-Z0-9]*LOT[:\s]*([A-Z0-9\-]+)", re.IGNORECASE)


def canonicalize_sku(raw: str) -> str | None:
    s = (raw or "").upper().strip()
    # Exclude IFUs and non-device items
    if s in EXCLUDED_SKUS or s.upper() in [x.upper() for x in EXCLUDED_SKUS]:
        return None
    if s in VALID_SKUS:
        return s
    if "14" in s:
        return "211410SPT"
    if "16" in s:
        return "211610SPT"
    if "18" in s:
        return "211810SPT"
    return None


def normalize_lot(code: str) -> str:
    """
    Normalize lot to always have SLQ- prefix (legacy behavior).
    - Uppercase + strip
    - SLQ123 -> SLQ-123
    - 05012025 -> SLQ-05012025
    """
    c = (code or "").strip().upper()
    if not c:
        return ""
    # Already has SLQ- prefix
    if c.startswith("SLQ-"):
        return c
    # Has SLQ but no dash (SLQ12345 -> SLQ-12345)
    if c.startswith("SLQ"):
        return "SLQ-" + c[3:].lstrip("-")
    # Bare number or other code -> prefix with SLQ-
    return "SLQ-" + c


def extract_lot(text: str) -> str | None:
    """
    Lean lot heuristic:
    - Prefer explicit "LOT: <code>".
    - Otherwise, look for SLQ-12345 / SLQ12345 patterns.
    - Fall back to bare numeric codes (6-12 digits).
    """
    t = (text or "").strip()
    if not t:
        return None
    # 1) Explicit LOT: label
    m = LOT_LABEL_RX.search(t)
    if m:
        lot = normalize_lot(m.group(1))
        return lot or None
    # 2) SLQ pattern
    m2 = LOT_RX.search(t)
    if m2:
        lot = normalize_lot(m2.group(0))
        return lot or None
    # 3) Bare numeric (e.g., "05012025")
    m3 = BARE_LOT_RX.search(t)
    if m3:
        lot = normalize_lot(m3.group(1))
        return lot or None
    return None


def extract_sku_lot_pairs(text: str) -> dict[str, str]:
    """
    Extract multiple SKU→LOT pairs from internal notes.
    
    Example input: "SKU: 21600101003 lot: SLQ-05012025 SKU: 21800101003 LOT: SLQ-05022025"
    Returns: {"211610SPT": "SLQ-05012025", "211810SPT": "SLQ-05022025"}
    """
    t = (text or "").strip()
    if not t:
        return {}
    
    pairs: dict[str, str] = {}
    for match in SKU_LOT_PAIR_RX.finditer(t):
        raw_sku = match.group(1)
        raw_lot = match.group(2)
        canonical_sku = canonicalize_sku(raw_sku)
        if canonical_sku:
            normalized_lot = normalize_lot(raw_lot)
            if normalized_lot:
                pairs[canonical_sku] = normalized_lot
    
    return pairs


def infer_units(item_name: str, quantity: int) -> int:
    """
    Convert ordered quantity to individual units.

    ShipStation item names may indicate:
    - "Box of 10" / "10-pack" -> multiply by 10
    - "Case of 100" -> multiply by 100
    - Individual units -> return as-is
    """
    name = (item_name or "").lower()
    qty = int(quantity or 0)
    if qty <= 0:
        return 0
    box_10_patterns = [
        "10-pack", "10 pack", "10pk", "box of 10", "bx of 10",
        "pack of 10", "pk of 10", "10/box", "10/pk",
    ]
    for pattern in box_10_patterns:
        if pattern in name:
            return qty * 10
    if "box of 5" in name or "5-pack" in name or "5 pack" in name:
        return qty * 5
    if "case of 100" in name or "100/case" in name:
        return qty * 100
    return qty


def _read_lotlog_bytes(path_str: str) -> bytes | None:
    """
    Read LotLog.csv bytes, trying:
    1. Storage backend (key = "data/LotLog.csv")
    2. Local file at the given path
    """
    # Try storage backend first
    try:
        from flask import current_app
        from app.eqms.storage import storage_from_config
        storage = storage_from_config(current_app.config)
        if storage.exists("data/LotLog.csv"):
            return storage.get_bytes("data/LotLog.csv")
    except Exception:
        pass  # Not in Flask context or storage not configured

    # Fall back to local file
    p = Path(path_str.replace("\\", "/"))
    if p.exists():
        return p.read_bytes()

    return None


def load_lot_log(path_str: str) -> tuple[dict[str, str], dict[str, str]]:
    """
    Load LotLog.csv mapping:
    - lot_to_sku: {lot_variant -> canonical_sku}
    - lot_corrections: {raw_lot -> correct_lot} (from "Correct Lot Name" column)
    
    Stores multiple variants for reliable lookup:
    - Normalized lot (SLQ-05012025)
    - Without prefix (05012025)
    - Raw uppercase
    """
    raw_bytes = _read_lotlog_bytes(path_str)
    if not raw_bytes:
        return {}, {}
    
    lot_to_sku: dict[str, str] = {}
    lot_corrections: dict[str, str] = {}

    import io
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        raw_lot = (str(row.get("Lot") or "")).strip().upper()
        correct_lot_name = (str(row.get("Correct Lot Name") or "")).strip().upper()
        sku_raw = str(row.get("SKU") or "")
        sku = canonicalize_sku(sku_raw)

        if not raw_lot or not sku:
            continue

        # Determine the canonical lot (prefer "Correct Lot Name" if present)
        if correct_lot_name:
            canonical_lot = normalize_lot(correct_lot_name)
            # Store correction mapping
            norm_raw = normalize_lot(raw_lot)
            if norm_raw != canonical_lot:
                lot_corrections[norm_raw] = canonical_lot
                lot_corrections[raw_lot] = canonical_lot
        else:
            canonical_lot = normalize_lot(raw_lot)

        # Store multiple variants -> SKU
        lot_to_sku[canonical_lot] = sku
        lot_to_sku[raw_lot] = sku
        lot_to_sku[normalize_lot(raw_lot)] = sku

        # Store without SLQ- prefix
        if canonical_lot.startswith("SLQ-"):
            lot_to_sku[canonical_lot[4:]] = sku
        if raw_lot.startswith("SLQ-"):
            lot_to_sku[raw_lot[4:]] = sku

    return lot_to_sku, lot_corrections


def load_lot_log_with_inventory(path_str: str) -> tuple[dict[str, str], dict[str, str], dict[str, int], dict[str, int]]:
    """
    Load LotLog.csv with inventory data:
    - lot_to_sku: {lot_variant -> canonical_sku}
    - lot_corrections: {raw_lot -> correct_lot}
    - lot_inventory: {canonical_lot -> total_units_produced}
    - lot_years: {canonical_lot -> manufacturing_year}
    """
    raw_bytes = _read_lotlog_bytes(path_str)
    if not raw_bytes:
        return {}, {}, {}, {}

    lot_to_sku: dict[str, str] = {}
    lot_corrections: dict[str, str] = {}
    lot_inventory: dict[str, int] = {}
    lot_years: dict[str, int] = {}

    import io
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        raw_lot = (str(row.get("Lot") or "")).strip().upper()
        correct_lot_name = (str(row.get("Correct Lot Name") or "")).strip().upper()
        sku_raw = str(row.get("SKU") or "")
        sku = canonicalize_sku(sku_raw)

        if not raw_lot or not sku:
            continue

        # Determine canonical lot (prefer "Correct Lot Name")
        if correct_lot_name:
            canonical_lot = normalize_lot(correct_lot_name)
            norm_raw = normalize_lot(raw_lot)
            if norm_raw != canonical_lot:
                lot_corrections[norm_raw] = canonical_lot
                lot_corrections[raw_lot] = canonical_lot
        else:
            canonical_lot = normalize_lot(raw_lot)

        # Store inventory (Total Units in Lot)
        try:
            total_units = int(float(row.get("Total Units in Lot") or 0))
        except Exception:
            total_units = 0
        if canonical_lot:
            lot_inventory[canonical_lot] = total_units

        # Manufacturing year (from Lot Log or lot string)
        mfg_date = (str(row.get("Manufacturing Date") or "")).strip()
        year_val = None
        if mfg_date:
            m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", mfg_date)
            if m:
                try:
                    year_val = int(m.group(3))
                except Exception:
                    year_val = None
            if not year_val:
                m = re.match(r"(\d{4})-\d{2}-\d{2}", mfg_date)
                if m:
                    try:
                        year_val = int(m.group(1))
                    except Exception:
                        year_val = None
            if not year_val:
                try:
                    year_val = int(mfg_date[:4])
                except Exception:
                    year_val = None
        if not year_val:
            m = re.search(r"(20\d{2})", canonical_lot)
            if m:
                try:
                    year_val = int(m.group(1))
                except Exception:
                    year_val = None
            if not year_val:
                digits = re.sub(r"\D", "", canonical_lot or "")
                if len(digits) >= 4:
                    try:
                        candidate = int(digits[-4:])
                        if 2000 <= candidate <= 2100:
                            year_val = candidate
                    except Exception:
                        year_val = None
        if year_val:
            lot_years[canonical_lot] = year_val

        # Store multiple variants -> SKU
        lot_to_sku[canonical_lot] = sku
        lot_to_sku[raw_lot] = sku
        lot_to_sku[normalize_lot(raw_lot)] = sku
        if canonical_lot.startswith("SLQ-"):
            lot_to_sku[canonical_lot[4:]] = sku
        if raw_lot.startswith("SLQ-"):
            lot_to_sku[raw_lot[4:]] = sku

    return lot_to_sku, lot_corrections, lot_inventory, lot_years


def load_lot_dates(path_str: str) -> tuple[dict[str, str], dict[str, str]]:
    """
    Load manufacturing and expiration dates per canonical lot from LotLog.csv.
    Returns (lot_mfg_dates, lot_exp_dates) where keys are canonical lot names
    and values are date strings (as-is from CSV).
    """
    raw_bytes = _read_lotlog_bytes(path_str)
    if not raw_bytes:
        return {}, {}

    lot_mfg_dates: dict[str, str] = {}
    lot_exp_dates: dict[str, str] = {}

    import io
    text = raw_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        correct_lot = (str(row.get("Correct Lot Name") or "")).strip().upper()
        raw_lot = (str(row.get("Lot") or "")).strip().upper()

        canonical = normalize_lot(correct_lot) if correct_lot else normalize_lot(raw_lot)
        if not canonical:
            continue

        mfg = (str(row.get("Manufacturing Date") or "")).strip()
        exp = (str(row.get("Expiration Date") or row.get("Exp Date") or "")).strip()

        if mfg and canonical not in lot_mfg_dates:
            lot_mfg_dates[canonical] = mfg
        if exp and canonical not in lot_exp_dates:
            lot_exp_dates[canonical] = exp

    return lot_mfg_dates, lot_exp_dates


def resolve_disposition_log_path() -> str:
    """Absolute path to default DispositionLog.xlsx (local fallback)."""
    env_path = (os.environ.get("DISPOSITION_LOG_PATH") or "").strip()
    if env_path:
        return env_path
    project_root = Path(__file__).resolve().parents[4]
    return str(project_root / "app" / "eqms" / "data" / "DispositionLog.xlsx")


def _read_disposition_log_bytes(path_str: str) -> bytes | None:
    """Read DispositionLog.xlsx from storage backend or local path."""
    try:
        from flask import current_app
        from app.eqms.storage import storage_from_config

        storage = storage_from_config(current_app.config)
        if storage.exists("data/DispositionLog.xlsx"):
            return storage.get_bytes("data/DispositionLog.xlsx")
    except Exception:
        pass

    p = Path(path_str.replace("\\", "/"))
    if p.exists():
        return p.read_bytes()
    return None


def parse_disposition_log_bytes(
    file_bytes: bytes,
    *,
    lot_corrections: dict[str, str] | None = None,
) -> dict[str, int]:
    """
    Parse DispositionLog.xlsx rows into {canonical_lot: total_units_dispositioned}.

    Expected columns: Date, Lot, SKU, Number of Units Dispositioned
    """
    import io

    from openpyxl import load_workbook

    lot_corrections = lot_corrections or {}
    totals: dict[str, int] = {}
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return totals
        headers = [str(h or "").strip() for h in header]
        col = {name.lower(): idx for idx, name in enumerate(headers)}
        lot_idx = col.get("lot")
        qty_idx = col.get("number of units dispositioned")
        if lot_idx is None or qty_idx is None:
            return totals
        for row in rows:
            if not row or lot_idx >= len(row):
                continue
            raw_lot = str(row[lot_idx] or "").strip()
            if not raw_lot:
                continue
            try:
                qty = int(float(row[qty_idx] or 0))
            except (TypeError, ValueError):
                qty = 0
            if qty <= 0:
                continue
            normalized = normalize_lot(raw_lot)
            corrected = lot_corrections.get(normalized, normalized)
            totals[corrected] = totals.get(corrected, 0) + qty
    finally:
        wb.close()
    return totals


def load_disposition_log(
    path_str: str,
    *,
    lot_corrections: dict[str, str] | None = None,
) -> dict[str, int]:
    """Load disposition totals per canonical lot from storage or local xlsx."""
    raw_bytes = _read_disposition_log_bytes(path_str)
    if not raw_bytes:
        return {}
    return parse_disposition_log_bytes(raw_bytes, lot_corrections=lot_corrections)

