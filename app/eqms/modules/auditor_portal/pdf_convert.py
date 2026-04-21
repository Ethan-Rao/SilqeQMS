from __future__ import annotations

import hashlib
import io
import logging
import os
import shutil
import subprocess
import tempfile
from typing import Any

from flask import current_app

from app.eqms.document_viewer import sanitize_viewer_html
from app.eqms.storage import StorageError, storage_from_config

logger = logging.getLogger(__name__)

_EXCEL_ROW_LIMIT = 2000
_SOFFICE_TIMEOUT_SEC = 120


def _soffice_binary() -> str | None:
    """Return the path to the soffice binary if available, else None."""
    override = os.environ.get("SOFFICE_BIN")
    if override and os.path.isfile(override):
        return override
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path:
            return path
    return None


def soffice_convert_to_pdf(file_bytes: bytes, source_ext: str) -> bytes:
    """Convert Word / Excel bytes to PDF using headless LibreOffice.

    Uses a per-invocation user profile directory so concurrent workers don't
    deadlock on the shared profile lock. Raises RuntimeError on any failure
    so callers can fall back to xhtml2pdf.
    """
    soffice = _soffice_binary()
    if not soffice:
        raise RuntimeError("soffice binary not found on PATH (install libreoffice-writer + libreoffice-calc)")

    ext = source_ext.lower().lstrip(".")
    if ext not in {"docx", "doc", "xlsx", "xls", "csv", "odt", "ods"}:
        raise RuntimeError(f"unsupported soffice source extension: {source_ext!r}")

    with tempfile.TemporaryDirectory(prefix="lo_") as tmpdir:
        profile_dir = os.path.join(tmpdir, "profile")
        os.makedirs(profile_dir, exist_ok=True)
        input_path = os.path.join(tmpdir, f"in.{ext}")
        with open(input_path, "wb") as f:
            f.write(file_bytes)

        cmd = [
            soffice,
            f"-env:UserInstallation=file://{profile_dir}",
            "--headless",
            "--nologo",
            "--nofirststartwizard",
            "--norestore",
            "--convert-to", "pdf",
            "--outdir", tmpdir,
            input_path,
        ]
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                timeout=_SOFFICE_TIMEOUT_SEC,
                check=False,
            )
        except subprocess.TimeoutExpired as e:
            raise RuntimeError(f"soffice timed out after {_SOFFICE_TIMEOUT_SEC}s") from e

        if result.returncode != 0:
            stderr_tail = (result.stderr or b"")[-500:].decode("utf-8", errors="replace")
            stdout_tail = (result.stdout or b"")[-500:].decode("utf-8", errors="replace")
            raise RuntimeError(
                f"soffice exited {result.returncode}; stderr={stderr_tail!r} stdout={stdout_tail!r}"
            )

        pdf_path = os.path.join(tmpdir, f"in.pdf")
        if not os.path.isfile(pdf_path):
            stderr_tail = (result.stderr or b"")[-500:].decode("utf-8", errors="replace")
            raise RuntimeError(f"soffice produced no PDF; stderr={stderr_tail!r}")

        with open(pdf_path, "rb") as f:
            data = f.read()
        if not data.startswith(b"%PDF"):
            raise RuntimeError("soffice output does not look like a PDF")
        return data


def active_backend() -> str:
    """Resolve the effective PDF backend.

    'auto' (the default) tries LibreOffice first and falls back to xhtml2pdf
    when soffice is missing (local dev without LO installed).
    """
    v = (current_app.config.get("AUDITOR_PDF_BACKEND") or "auto").strip().lower()
    if v not in {"auto", "libreoffice", "xhtml2pdf"}:
        v = "auto"
    if v == "auto":
        return "libreoffice" if _soffice_binary() else "xhtml2pdf"
    return v


def render_excel_to_html(file_bytes: bytes) -> str | None:
    """Return standalone HTML (body content suitable for wrapping) for Excel files."""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return None
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return None
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_raw: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows_raw.append([str(c) if c is not None else "" for c in row])
            total = max(0, len(rows_raw) - 1)
            truncated = False
            if len(rows_raw) > _EXCEL_ROW_LIMIT + 1:
                rows_raw = rows_raw[: _EXCEL_ROW_LIMIT + 1]
                truncated = True
            if not rows_raw:
                continue
            banner = ""
            if truncated:
                banner = (
                    f'<div class="banner">Showing first {_EXCEL_ROW_LIMIT} of {total} data rows '
                    f"on sheet {_esc(sheet_name)}.</div>"
                )
            thead = "<thead><tr>" + "".join(f"<th>{_esc(c)}</th>" for c in rows_raw[0]) + "</tr></thead>"
            tbody_rows = rows_raw[1:] if len(rows_raw) > 1 else []
            tbody = "<tbody>" + "".join(
                "<tr>" + "".join(f"<td>{_esc(c)}</td>" for c in r) + "</tr>" for r in tbody_rows
            ) + "</tbody>"
            parts.append(
                f'<section class="sheet" style="page-break-after:always;">'
                f"<h2>{_esc(sheet_name)}</h2>{banner}"
                f'<table border="1" cellpadding="4" cellspacing="0">{thead}{tbody}</table></section>'
            )
        wb.close()
    except Exception:
        return None
    if not parts:
        return None
    return (
        "<!DOCTYPE html><html><head><meta charset='utf-8'/>"
        "<style>body{font-family:sans-serif;font-size:10pt;} table{border-collapse:collapse;width:100%;}"
        ".banner{background:#fee;padding:8px;margin:8px 0;} h2{font-size:12pt;}</style></head><body>"
        + "".join(parts)
        + "</body></html>"
    )


def _esc(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def html_to_pdf_bytes(html: str, base_url: str | None = None) -> bytes:
    """HTML -> PDF via xhtml2pdf. Kept for the fallback path only."""
    buf = io.BytesIO()
    try:
        from xhtml2pdf import pisa  # type: ignore
    except ImportError as e:
        raise RuntimeError("xhtml2pdf not installed") from e
    pisa.CreatePDF(io.StringIO(html), dest=buf, encoding="utf-8")
    return buf.getvalue()


def _docx_via_xhtml2pdf(file_bytes: bytes) -> bytes:
    try:
        import mammoth  # type: ignore
    except ImportError as e:
        raise RuntimeError("mammoth not installed") from e
    result = mammoth.convert_to_html(io.BytesIO(file_bytes))
    body = sanitize_viewer_html(result.value)
    shell = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'/>"
        "<style>body{font-family:sans-serif;font-size:11pt;} table{border-collapse:collapse;}"
        "td,th{border:1px solid #ccc;padding:4px;}</style></head><body>"
        f"{body}</body></html>"
    )
    return html_to_pdf_bytes(shell)


def _xlsx_via_xhtml2pdf(file_bytes: bytes) -> bytes:
    html = render_excel_to_html(file_bytes)
    if not html:
        raise RuntimeError("excel html failed")
    return html_to_pdf_bytes(html)


def docx_bytes_to_pdf(file_bytes: bytes) -> bytes:
    """docx -> PDF. Prefers LibreOffice for Word-identical layout; falls back
    to the mammoth + xhtml2pdf pipeline if LO is unavailable or fails."""
    backend = active_backend()
    if backend == "libreoffice":
        try:
            return soffice_convert_to_pdf(file_bytes, ".docx")
        except Exception as e:
            logger.warning("LibreOffice docx conversion failed (%s); falling back to xhtml2pdf", e)
    return _docx_via_xhtml2pdf(file_bytes)


def xlsx_bytes_to_pdf(file_bytes: bytes) -> bytes:
    """xlsx -> PDF. Prefers LibreOffice so column widths, sheets, and cell
    formatting match what Excel would print; falls back to HTML-table PDF."""
    backend = active_backend()
    if backend == "libreoffice":
        try:
            return soffice_convert_to_pdf(file_bytes, ".xlsx")
        except Exception as e:
            logger.warning("LibreOffice xlsx conversion failed (%s); falling back to xhtml2pdf", e)
    return _xlsx_via_xhtml2pdf(file_bytes)


class CachedPdfStore:
    def __init__(self, config: dict[str, Any]) -> None:
        self._storage = storage_from_config(config)

    def _full_key(self, namespace: str, cache_key: str) -> str:
        # Include backend in the key so a LibreOffice-rendered PDF never
        # serves from an older xhtml2pdf cache entry (and vice versa).
        try:
            backend = active_backend()
        except Exception:
            backend = "unknown"
        return f"auditor-cache/{namespace}/{backend}/{cache_key}.pdf"

    def get(self, namespace: str, cache_key: str) -> bytes | None:
        key = self._full_key(namespace, cache_key)
        try:
            if hasattr(self._storage, "exists") and not self._storage.exists(key):
                return None
            return self._storage.get_bytes(key)
        except Exception as e:
            logger.debug("auditor cache miss/read error: %s", e)
            return None

    def put(self, namespace: str, cache_key: str, data: bytes) -> None:
        key = self._full_key(namespace, cache_key)
        try:
            self._storage.put_bytes(key, data, content_type="application/pdf")
        except StorageError:
            logger.warning("auditor cache write failed for key=%s", key)
        except Exception as e:
            logger.warning("auditor cache write failed: %s", e)
