from __future__ import annotations

import hashlib
from datetime import date, datetime, timedelta
from email import policy
from email.message import Message
from email.parser import BytesParser
from typing import TYPE_CHECKING

from werkzeug.utils import secure_filename

from app.eqms.utils import utcnow

from app.eqms.audit import record_event

if TYPE_CHECKING:
    from sqlalchemy.orm import Session
    from app.eqms.models import User
    from app.eqms.modules.purchasing.models import PurchaseOrder, PurchaseOrderAttachment, PurchaseOrderLine


def parse_date(s: str | None) -> date | None:
    """Parse YYYY-MM-DD date string."""
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return None


def validate_purchase_order_payload(payload: dict) -> list[str]:
    errors = []
    if not (payload.get("po_number") or "").strip():
        errors.append("PO number is required.")
    if not payload.get("order_date"):
        errors.append("Order date is required.")
    return errors


PO_ATTACHMENT_TYPES = (
    "po_pdf",
    "confirmation_pdf",
    "confirmation_eml",
    "other",
    "verification_evidence",
)


def apply_supplier_choice(payload: dict) -> dict:
    """D58: approved-supplier select wins; otherwise store free-text and leave supplier_id null."""
    sid = payload.get("supplier_id")
    if sid:
        payload["supplier_id"] = int(sid)
        payload["supplier_name"] = None
    else:
        payload["supplier_id"] = None
        payload["supplier_name"] = (payload.get("supplier_name") or "").strip() or None
    return payload


def _digest(file_bytes: bytes) -> tuple[str, int]:
    h = hashlib.sha256()
    h.update(file_bytes)
    return h.hexdigest(), len(file_bytes)


def build_po_storage_key(po_number: str, filename: str, upload_date: date | None = None) -> str:
    if upload_date is None:
        upload_date = date.today()
    safe_po = po_number.replace("/", "_").replace("\\", "_")
    safe_filename = secure_filename(filename) or "document.bin"
    return f"purchase_orders/{safe_po}/{upload_date.isoformat()}/{safe_filename}"


def parse_line_items(items_text: str | None) -> list[dict]:
    """Parse line items from newline-delimited text."""
    lines = []
    for raw in (items_text or "").splitlines():
        row = raw.strip()
        if not row:
            continue
        # Preferred: item_code | description | qty | unit_price
        if "|" in row:
            parts = [p.strip() for p in row.split("|")]
            item_code = parts[0] or None
            description = parts[1] if len(parts) > 1 else None
            qty = parts[2] if len(parts) > 2 else None
            unit_price = parts[3] if len(parts) > 3 else None
            lines.append(
                {
                    "item_code": item_code,
                    "description": description,
                    "quantity": int(qty) if qty and qty.isdigit() else 1,
                    "unit_price": unit_price,
                }
            )
            continue
        # Accept "qty x description" or "qty, description"
        qty = 1
        description = row
        if " x " in row.lower():
            parts = row.lower().split(" x ", 1)
            if parts[0].strip().isdigit():
                qty = int(parts[0].strip())
                description = row[len(parts[0]) + 3 :].strip()
        elif "," in row:
            parts = row.split(",", 1)
            if parts[0].strip().isdigit():
                qty = int(parts[0].strip())
                description = parts[1].strip()
        lines.append({"description": description or row, "quantity": qty})
    return lines


def create_purchase_order(s: "Session", payload: dict, user: "User") -> "PurchaseOrder":
    from app.eqms.modules.purchasing.models import PurchaseOrder, PurchaseOrderLine

    now = utcnow()
    po = PurchaseOrder(
        po_number=(payload.get("po_number") or "").strip(),
        order_date=payload.get("order_date"),
        expected_date=payload.get("expected_date"),
        received_date=payload.get("received_date"),
        payment_due_date=payload.get("payment_due_date"),
        supplier_id=payload.get("supplier_id"),
        supplier_name=(
            None
            if payload.get("supplier_id")
            else ((payload.get("supplier_name") or "").strip() or None)
        ),
        status=(payload.get("status") or "pending").strip(),
        description=(payload.get("description") or "").strip() or None,
        notes=(payload.get("notes") or "").strip() or None,
        amount=(payload.get("amount") or "").strip() or None,
        meets_requirements=(payload.get("meets_requirements") or "").strip() or None,
        verified_how=(payload.get("verified_how") or "").strip() or None,
        closed_by=(payload.get("closed_by") or "").strip() or None,
        reference=(payload.get("reference") or "").strip() or None,
        created_at=now,
        updated_at=now,
        created_by_user_id=user.id,
    )
    s.add(po)
    s.flush()

    for line in payload.get("lines") or []:
        item = PurchaseOrderLine(
            purchase_order_id=po.id,
            item_code=(line.get("item_code") or "").strip() or None,
            description=(line.get("description") or "").strip() or None,
            quantity=int(line.get("quantity") or 1),
            quantity_received=int(line.get("quantity_received") or 0),
            unit_price=(line.get("unit_price") or "").strip() or None,
        )
        s.add(item)

    record_event(
        s,
        actor=user,
        action="purchase_order.create",
        entity_type="PurchaseOrder",
        entity_id=str(po.id),
        metadata={"po_number": po.po_number},
    )
    return po


def update_purchase_order(s: "Session", po: "PurchaseOrder", payload: dict, user: "User", reason: str | None = None) -> "PurchaseOrder":
    from datetime import datetime as _dt
    from decimal import Decimal as _Decimal

    def _audit_val(val):
        if isinstance(val, _dt):
            return val.isoformat()
        if isinstance(val, date):
            return val.isoformat()
        if isinstance(val, _Decimal):
            return str(val)
        return val

    changes = {}

    def _set(attr: str, val):
        nonlocal changes
        if val != getattr(po, attr):
            changes[attr] = {"old": _audit_val(getattr(po, attr)), "new": _audit_val(val)}
            setattr(po, attr, val)

    _set("order_date", payload.get("order_date") or po.order_date)
    _set("expected_date", payload.get("expected_date"))
    _set("received_date", payload.get("received_date"))
    _set("payment_due_date", payload.get("payment_due_date"))
    _set("supplier_id", payload.get("supplier_id"))
    if "supplier_name" in payload or "supplier_id" in payload:
        if payload.get("supplier_id"):
            _set("supplier_name", None)
        else:
            _set("supplier_name", (payload.get("supplier_name") or "").strip() or None)
    _set("status", (payload.get("status") or po.status).strip())
    _set("description", (payload.get("description") or "").strip() or None)
    _set("notes", (payload.get("notes") or "").strip() or None)
    _set("amount", (payload.get("amount") or "").strip() or None)
    _set("meets_requirements", (payload.get("meets_requirements") or "").strip() or None)
    _set("verified_how", (payload.get("verified_how") or "").strip() or None)
    _set("closed_by", (payload.get("closed_by") or "").strip() or None)
    _set("reference", (payload.get("reference") or "").strip() or None)

    if "is_closed" in payload:
        want_closed = bool(payload.get("is_closed"))
        _set("is_closed", want_closed)
        if want_closed:
            if payload.get("closed_at") is not None:
                _set("closed_at", payload.get("closed_at"))
            elif po.closed_at is None:
                _set("closed_at", date.today())
        else:
            _set("closed_at", None)

    po.updated_at = utcnow()

    record_event(
        s,
        actor=user,
        action="purchase_order.edit",
        entity_type="PurchaseOrder",
        entity_id=str(po.id),
        reason=reason,
        metadata={"changes": changes},
    )
    return po


def upload_purchase_order_attachment(
    s: "Session",
    po: "PurchaseOrder",
    file_bytes: bytes,
    filename: str,
    content_type: str,
    user: "User",
    attachment_type: str,
) -> "PurchaseOrderAttachment":
    from flask import current_app
    from app.eqms.storage import storage_from_config
    from app.eqms.modules.purchasing.models import PurchaseOrderAttachment

    sha256, size_bytes = _digest(file_bytes)
    storage_key = build_po_storage_key(po.po_number, filename)

    storage = storage_from_config(current_app.config)
    storage.put_bytes(storage_key, file_bytes, content_type=content_type)

    attachment = PurchaseOrderAttachment(
        purchase_order_id=po.id,
        attachment_type=attachment_type,
        storage_key=storage_key,
        filename=secure_filename(filename) or "document.bin",
        content_type=content_type,
        size_bytes=size_bytes,
        uploaded_by_user_id=user.id,
    )
    s.add(attachment)
    s.flush()

    record_event(
        s,
        actor=user,
        action="purchase_order.attachment_upload",
        entity_type="PurchaseOrderAttachment",
        entity_id=str(attachment.id),
        metadata={"po_id": po.id, "filename": attachment.filename, "type": attachment_type, "sha256": sha256},
    )
    return attachment


_EXCEL_EPOCH = date(1899, 12, 30)


def coerce_po_date(value) -> date | None:
    """
    Coerce a PO Log cell into a date. Handles datetime/date objects, Excel serial
    numbers (e.g. 44631), and human strings ("05 Mar 2024", "3 March 2025",
    "2019-10-24 00:00:00"). Returns None for blanks / "N/A".
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date. Plausible range only (avoids treating stray numbers as dates).
        if 20000 <= value <= 60000:
            return _EXCEL_EPOCH + timedelta(days=int(value))
        return None
    text = str(value).strip()
    if not text or text.upper() in ("N/A", "NA", "NONE", "-"):
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        pass
    for fmt in (
        "%m/%d/%Y", "%m/%d/%y", "%d %b %Y", "%d%b%Y", "%d %B %Y",
        "%b %d %Y", "%B %d %Y", "%d %B, %Y", "%B %d, %Y",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    # Excel serial embedded as string
    if text.isdigit() and 20000 <= int(text) <= 60000:
        return _EXCEL_EPOCH + timedelta(days=int(text))
    return None


def _po_cell_text(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


SUPPLIER_FROM_PO_LOG_PREFIX = "Supplier from PO Log: "

PO_LOG_HEADERS = [
    "P.O. Number",
    "Supplier/Vendor Name and identification",
    "Date",
    "Target Delivery Date",
    "Actual Delivery Date",
    "Product/Service Meets Requirement(s)?\nYes / No",
    "Verified how?",
    "Closed by\nInitials / Date",
    "Cost Info.",
    "References",
    "Notes/Comments",
]


def _is_blank(value) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def apply_po_blank_fills(po, values: dict) -> list[str]:
    """Set only currently blank attributes. Returns list of field names filled."""
    filled: list[str] = []
    for attr, val in values.items():
        if val is None:
            continue
        if isinstance(val, str) and not val.strip():
            continue
        if not _is_blank(getattr(po, attr, None)):
            continue
        setattr(po, attr, val)
        filled.append(attr)
    return filled


def resolve_supplier_by_extracted_name(s: "Session", name: str | None) -> tuple[object | None, str]:
    """Match extracted supplier text to existing Supplier rows (D53).

    Uses ``canonical_customer_key`` (corporate-suffix stripping) rather than bare ilike.
    Returns ``(supplier_or_None, status)`` where status is ``unique``, ``none``, or ``ambiguous``.
    Never creates a Supplier row.
    """
    from app.eqms.modules.customer_profiles.utils import canonical_customer_key
    from app.eqms.modules.suppliers.models import Supplier

    raw = (name or "").strip()
    if not raw:
        return None, "none"
    key = canonical_customer_key(raw)
    if not key:
        return None, "none"

    matches: list = []
    for sup in s.query(Supplier).all():
        if canonical_customer_key(sup.name or "") == key:
            matches.append(sup)
    if len(matches) == 1:
        return matches[0], "unique"
    if len(matches) > 1:
        return None, "ambiguous"
    return None, "none"


def append_po_lines_if_empty(s: "Session", po, items: list[dict] | None) -> str | None:
    """Write extracted lines only when the PO has none (D52). Returns a status token."""
    from app.eqms.modules.purchasing.models import PurchaseOrderLine

    existing = list(po.lines or [])
    extracted = list(items or [])
    if not extracted:
        return None
    if existing:
        return "lines_skipped_existing"
    for line in extracted:
        s.add(
            PurchaseOrderLine(
                purchase_order_id=po.id,
                item_code=(line.get("item_code") or "").strip() or None,
                description=(line.get("description") or "").strip() or None,
                quantity=int(line.get("quantity") or 1),
                quantity_received=int(line.get("quantity_received") or 0),
                unit_price=(str(line.get("unit_price")).strip() if line.get("unit_price") is not None else None)
                or None,
            )
        )
    return "lines_added"


def stage_po_pdf_bytes(file_bytes: bytes, filename: str, content_type: str = "application/pdf") -> dict:
    """Stage upload bytes under temp_po_pdf/ for review-form round trips (D50)."""
    import uuid

    from flask import current_app
    from app.eqms.storage import storage_from_config

    sha256, size_bytes = _digest(file_bytes)
    ref = f"{sha256[:16]}_{uuid.uuid4().hex[:10]}"
    storage_key = f"temp_po_pdf/{ref}.pdf"
    storage = storage_from_config(current_app.config)
    storage.put_bytes(storage_key, file_bytes, content_type=content_type or "application/pdf")
    return {
        "storage_key": storage_key,
        "filename": secure_filename(filename) or "document.pdf",
        "content_type": content_type or "application/pdf",
        "size_bytes": size_bytes,
        "sha256": sha256,
    }


def delete_staged_po_pdf(storage_key: str | None) -> None:
    if not storage_key or not str(storage_key).startswith("temp_po_pdf/"):
        return
    from flask import current_app
    from app.eqms.storage import storage_from_config

    storage = storage_from_config(current_app.config)
    try:
        storage.delete(storage_key)
    except Exception:
        pass


def cleanup_stale_temp_po_pdfs(*, max_age_hours: int = 24) -> int:
    """Best-effort purge of abandoned staging objects. Returns delete count.

    Called from the import page so abandoned reviews do not leave unreachable
    objects forever. Only deletes keys under ``temp_po_pdf/`` older than the age.
    """
    import time
    from flask import current_app
    from app.eqms.storage import storage_from_config

    storage = storage_from_config(current_app.config)
    deleted = 0
    list_fn = getattr(storage, "list_keys", None)
    mtime_fn = getattr(storage, "key_mtime", None)
    if not callable(list_fn) or not callable(mtime_fn):
        return 0
    try:
        keys = list_fn("temp_po_pdf/")
    except Exception:
        return 0
    cutoff = time.time() - (max_age_hours * 3600)
    for key in keys or []:
        if not str(key).startswith("temp_po_pdf/"):
            continue
        try:
            mtime = mtime_fn(key)
            if mtime is None or float(mtime) > cutoff:
                continue
            storage.delete(key)
            deleted += 1
        except Exception:
            continue
    return deleted


def supplier_name_for_export(po) -> str:
    if po.supplier and po.supplier.name:
        return po.supplier.name
    if (getattr(po, "supplier_name", None) or "").strip():
        return po.supplier_name.strip()
    notes = po.notes or ""
    if notes.startswith(SUPPLIER_FROM_PO_LOG_PREFIX):
        return notes[len(SUPPLIER_FROM_PO_LOG_PREFIX):].strip()
    return ""


def notes_for_export(po) -> str:
    notes = (po.notes or "").strip()
    if notes.startswith(SUPPLIER_FROM_PO_LOG_PREFIX):
        return ""
    return notes


def parse_closed_by_date(text: str | None) -> date | None:
    """Parse a date from closed_by text.

    Observed shapes: 'DP / 14Oct2022', 'ER 06Jan2025', 'ER/ 11Nov2024',
    and spaced 'ER 01 Feb 2026'.
    """
    import re

    if not text:
        return None
    m = re.search(r"(\d{1,2})([A-Za-z]{3})(\d{4})", text)
    if m:
        day, mon, year = m.group(1), m.group(2), m.group(3)
        try:
            return datetime.strptime(f"{int(day):02d}{mon}{year}", "%d%b%Y").date()
        except ValueError:
            pass
    m2 = re.search(r"(\d{1,2})\s+([A-Za-z]{3})\s+(\d{4})", text)
    if m2:
        try:
            return datetime.strptime(
                f"{int(m2.group(1)):02d} {m2.group(2)} {m2.group(3)}",
                "%d %b %Y",
            ).date()
        except ValueError:
            return None
    return None


def operator_initials(user) -> str:
    name = ((getattr(user, "display_name", None) or "") or (getattr(user, "email", None) or "")).strip()
    if not name:
        return "??"
    parts = [p for p in name.replace(".", " ").replace("_", " ").split() if p]
    if len(parts) >= 2:
        return (parts[0][0] + parts[-1][0]).upper()
    return (name[:2]).upper()


def format_closed_by(user, when: date | None = None) -> str:
    """House format matching 'DP / 14Oct2022'."""
    d = when or date.today()
    return f"{operator_initials(user)} / {d.strftime('%d%b%Y')}"


def _po_closure_snapshot(po) -> dict:
    return {
        "is_closed": bool(po.is_closed),
        "closed_at": po.closed_at.isoformat() if po.closed_at else None,
        "closed_by": po.closed_by,
        "status": po.status,
    }


def document_po_closed(s, *, po, user, when: date | None = None):
    """Set is_closed; fill closed_by only when blank. Never overwrite closed_by."""
    when = when or date.today()
    before = _po_closure_snapshot(po)
    po.is_closed = True
    po.closed_at = when
    if _is_blank(po.closed_by):
        po.closed_by = format_closed_by(user, when)
    po.updated_at = utcnow()
    after = _po_closure_snapshot(po)
    record_event(
        s,
        actor=user,
        action="purchase_order.closed",
        entity_type="PurchaseOrder",
        entity_id=str(po.id),
        metadata={"before": before, "after": after},
    )
    return po


def reopen_po(s, *, po, user):
    """Clear is_closed and closed_at; leave closed_by intact."""
    before = _po_closure_snapshot(po)
    po.is_closed = False
    po.closed_at = None
    po.updated_at = utcnow()
    after = _po_closure_snapshot(po)
    record_event(
        s,
        actor=user,
        action="purchase_order.reopened",
        entity_type="PurchaseOrder",
        entity_id=str(po.id),
        metadata={"before": before, "after": after},
    )
    return po


def build_po_log_xlsx(purchase_orders: list) -> bytes:
    """Build SILQ PO Log workbook matching import_po_log's expected layout (D37)."""
    import io

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Log"
    ws.append(["The P.O. Log process is defined in QM.SLQ020 Purchasing Controls SOP."])
    ws.append(["Obtain P.O. Number"])
    ws.append(list(PO_LOG_HEADERS))

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin

    for po in purchase_orders:
        row_idx = ws.max_row + 1
        values = [
            po.po_number,
            supplier_name_for_export(po),
            po.order_date,
            po.expected_date,
            po.received_date,
            po.meets_requirements or "",
            po.verified_how or "",
            po.closed_by or "",
            po.amount or "",
            po.reference or "",
            notes_for_export(po),
        ]
        ws.append(values)
        # Force PO number as text so Excel keeps leading zeros.
        po_cell = ws.cell(row=row_idx, column=1)
        po_cell.value = str(po.po_number)
        po_cell.number_format = "@"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_po_log(s: "Session", file_bytes: bytes, user: "User") -> dict:
    """
    Upsert purchase orders from the SILQ PO Log (.xlsx), keyed by P.O. Number.

    The header row is located wherever it appears; columns are mapped positionally
    from the "P.O. Number" column onward to match the SILQ log layout. Rows with no
    PO number or marked "*not used*" are skipped. Existing POs are updated in place;
    line items are never touched by this import.
    """
    import io
    import openpyxl

    from app.eqms.modules.purchasing.models import PurchaseOrder
    from app.eqms.modules.suppliers.models import Supplier

    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    # Fixed layout, offset from the "P.O. Number" column.
    # 0 PO#, 1 Supplier, 2 Order date, 3 Target delivery, 4 Actual delivery,
    # 5 Meets requirements, 6 Verified how, 7 Closed by, 8 Cost, 9 References, 10 Notes
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        header_idx = None
        po_col = None
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                if cell is None:
                    continue
                norm = " ".join(str(cell).split()).lower()
                # Match the real header cell ("P.O. Number") exactly so we don't trip on
                # the "Obtain P.O. Number" instruction row that precedes it.
                if norm in ("p.o. number", "po number", "p.o.number"):
                    header_idx = i
                    po_col = j
                    break
            if header_idx is not None:
                break
        if header_idx is None or po_col is None:
            continue

        for row in rows[header_idx + 1:]:
            def col(offset):
                idx = po_col + offset
                return row[idx] if idx < len(row) else None

            po_number = (_po_cell_text(col(0)) or "").strip()
            if not po_number:
                continue
            supplier_name = (_po_cell_text(col(1)) or "").strip()
            if not supplier_name or supplier_name.lower() in ("*not used*", "not used"):
                result["skipped"] += 1
                continue

            order_date = coerce_po_date(col(2))
            expected_date = coerce_po_date(col(3))
            received_date = coerce_po_date(col(4))
            meets = (_po_cell_text(col(5)) or "").strip() or None
            verified_how = (_po_cell_text(col(6)) or "").strip() or None
            closed_by = (_po_cell_text(col(7)) or "").strip() or None
            amount = (_po_cell_text(col(8)) or "").strip() or None
            reference = (_po_cell_text(col(9)) or "").strip() or None
            notes = (_po_cell_text(col(10)) or "").strip() or None

            supplier_id = None
            if supplier_name:
                supplier = (
                    s.query(Supplier).filter(Supplier.name.ilike(supplier_name)).first()
                )
                if supplier:
                    supplier_id = supplier.id
                elif not notes:
                    notes = f"Supplier from PO Log: {supplier_name}"

            status = "received" if received_date else "pending"

            try:
                existing = (
                    s.query(PurchaseOrder)
                    .filter(PurchaseOrder.po_number == po_number)
                    .one_or_none()
                )
                if existing:
                    # Fill blanks only (D38 / D2) — never overwrite operator or system data.
                    filled = apply_po_blank_fills(
                        existing,
                        {
                            "order_date": order_date,
                            "expected_date": expected_date,
                            "received_date": received_date,
                            "supplier_id": supplier_id,
                            "amount": amount,
                            "meets_requirements": meets,
                            "verified_how": verified_how,
                            "closed_by": closed_by,
                            "reference": reference,
                            "notes": notes,
                        },
                    )
                    if filled:
                        existing.updated_at = utcnow()
                        result["updated"] += 1
                    else:
                        result["skipped"] += 1
                else:
                    po = PurchaseOrder(
                        po_number=po_number,
                        order_date=order_date or date.today(),
                        expected_date=expected_date,
                        received_date=received_date,
                        supplier_id=supplier_id,
                        status=status,
                        notes=notes,
                        amount=amount,
                        meets_requirements=meets,
                        verified_how=verified_how,
                        closed_by=closed_by,
                        reference=reference,
                        created_at=utcnow(),
                        updated_at=utcnow(),
                        created_by_user_id=user.id,
                    )
                    s.add(po)
                    s.flush()
                    result["created"] += 1
            except Exception as e:  # noqa: BLE001 - collect per-row errors, keep importing
                result["errors"].append(f"{po_number}: {e}")
                result["skipped"] += 1

    wb.close()
    record_event(
        s,
        actor=user,
        action="purchase_order.import_log",
        entity_type="PurchaseOrder",
        entity_id="bulk",
        metadata={k: v for k, v in result.items() if k != "errors"},
    )
    return result


def _sanitize_eml_html(html: str) -> str:
    """Strip dangerous tags and event-handler attributes from EML HTML body (M-010)."""
    import re
    html = re.compile(
        r"<\s*/?\s*(script|iframe|object|embed|form|input|button|textarea|select|meta|link|base|applet)\b[^>]*>",
        re.IGNORECASE,
    ).sub("", html)
    html = re.compile(r"\s+on\w+\s*=", re.IGNORECASE).sub(" ", html)
    return html


def parse_eml_file(eml_bytes: bytes) -> dict:
    """Parse EML file and extract viewable content."""
    msg: Message = BytesParser(policy=policy.default).parsebytes(eml_bytes)
    result = {
        "from": msg.get("From", ""),
        "to": msg.get("To", ""),
        "cc": msg.get("Cc", ""),
        "subject": msg.get("Subject", ""),
        "date": msg.get("Date", ""),
        "body_text": "",
        "body_html": "",
        "attachments": [],
    }

    if msg.is_multipart():
        for part in msg.walk():
            content_type = part.get_content_type()
            if content_type == "text/plain" and not result["body_text"]:
                result["body_text"] = part.get_content()
            elif content_type == "text/html" and not result["body_html"]:
                result["body_html"] = _sanitize_eml_html(part.get_content())
            elif part.get_filename():
                result["attachments"].append(
                    {"filename": part.get_filename(), "content_type": content_type}
                )
    else:
        content_type = msg.get_content_type()
        if content_type == "text/plain":
            result["body_text"] = msg.get_content()
        elif content_type == "text/html":
            result["body_html"] = _sanitize_eml_html(msg.get_content())
    return result


# --------------------------------------------------------------------------- #
# P4-05 — Invoice upload / PO match / Other Payments
# --------------------------------------------------------------------------- #


class InvoiceFlowError(ValueError):
    """Operator-facing refusal for invoice flow actions."""


def _file_snapshot(atts) -> list[dict]:
    return [
        {
            "filename": a.filename,
            "storage_key": a.storage_key,
            "content_type": a.content_type,
            "size_bytes": a.size_bytes,
        }
        for a in (atts or [])
    ]


def _invoice_row_snapshot(entry) -> dict:
    return {
        "id": entry.id,
        "date_received": entry.date_received.isoformat() if entry.date_received else None,
        "payee": entry.payee,
        "description": entry.description,
        "amount": str(entry.amount) if entry.amount is not None else None,
        "due_date": entry.due_date.isoformat() if entry.due_date else None,
        "purchase_order_id": entry.purchase_order_id,
        "disposition": entry.disposition,
        "is_paid": bool(getattr(entry, "is_paid", False)),
        "files": _file_snapshot(entry.attachments),
    }


def _payment_row_snapshot(entry) -> dict:
    return {
        "id": entry.id,
        "vendor": entry.vendor,
        "description": entry.description,
        "amount": str(entry.amount) if entry.amount is not None else None,
        "payment_due_date": entry.payment_due_date.isoformat() if entry.payment_due_date else None,
        "invoice_received_entry_id": entry.invoice_received_entry_id,
        "files": _file_snapshot(entry.attachments),
        "line_item_files": [
            {
                "line_id": li.id,
                "description": li.description,
                "files": _file_snapshot(li.attachments),
            }
            for li in (entry.line_items or [])
        ],
    }


def migrate_payment_to_invoice(
    s,
    *,
    payment,
    file_bytes: bytes,
    filename: str,
    content_type: str | None,
    date_received: date | None,
    user,
    storage,
) -> "InvoiceReceivedEntry":
    """Upload invoice against an Upcoming payment: create received entry, move files.

    Never calls storage.delete — blobs are re-homed by storage_key only.
    """
    from werkzeug.utils import secure_filename

    from app.eqms.modules.purchasing.models import (
        InvoiceReceivedAttachment,
        InvoiceReceivedEntry,
        PaymentEntryAttachment,
    )

    if payment.invoice_received_entry_id:
        raise InvoiceFlowError("This payment has already been migrated to Invoices Received.")

    inv = InvoiceReceivedEntry(
        payee=payment.vendor,
        description=payment.description,
        amount=payment.amount,
        due_date=payment.payment_due_date,
        date_received=date_received or date.today(),
        disposition=InvoiceReceivedEntry.DISPOSITION_UNASSIGNED,
        created_by_id=user.id if user else None,
    )
    s.add(inv)
    s.flush()

    files_moved: list[str] = []

    # New uploaded invoice file
    safe_name = secure_filename(filename) or "invoice"
    storage_key = f"purchasing/invoice_received_files/{inv.id}/{safe_name}"
    storage.put_bytes(storage_key, file_bytes, content_type=content_type or "application/octet-stream")
    s.add(
        InvoiceReceivedAttachment(
            invoice_received_entry_id=inv.id,
            filename=filename,
            storage_key=storage_key,
            content_type=content_type or "application/octet-stream",
            size_bytes=len(file_bytes),
            uploaded_by_user_id=user.id if user else None,
        )
    )
    files_moved.append(filename)

    # Move payment-level attachments (row ownership only; do not touch Spaces).
    for att in list(payment.attachments or []):
        s.add(
            InvoiceReceivedAttachment(
                invoice_received_entry_id=inv.id,
                filename=att.filename,
                storage_key=att.storage_key,
                content_type=att.content_type,
                size_bytes=att.size_bytes,
                uploaded_by_user_id=att.uploaded_by_user_id,
            )
        )
        files_moved.append(att.filename)
        s.delete(att)
    payment.attachments = []

    # Move line-item attachments; keep the line item rows on the payment.
    for li in list(payment.line_items or []):
        for att in list(li.attachments or []):
            s.add(
                InvoiceReceivedAttachment(
                    invoice_received_entry_id=inv.id,
                    filename=att.filename,
                    storage_key=att.storage_key,
                    content_type=att.content_type,
                    size_bytes=att.size_bytes,
                    uploaded_by_user_id=att.uploaded_by_user_id,
                )
            )
            files_moved.append(att.filename)
            s.delete(att)
        li.attachments = []

    payment.invoice_received_entry_id = inv.id
    s.flush()

    record_event(
        s,
        actor=user,
        action="payment_entry.invoice_received",
        entity_type="PaymentEntry",
        entity_id=str(payment.id),
        metadata={
            "payment": _payment_row_snapshot(payment),
            "invoice": _invoice_row_snapshot(inv),
            "payment_entry_id": payment.id,
            "invoice_received_entry_id": inv.id,
            "files_moved": files_moved,
        },
    )
    return inv


def return_invoice_to_upcoming(s, *, invoice, user) -> "PaymentEntry":
    """Reverse migrate_payment_to_invoice. Never calls storage.delete."""
    from app.eqms.modules.purchasing.models import (
        PaymentEntry,
        PaymentEntryAttachment,
    )

    payment = (
        s.query(PaymentEntry)
        .filter(PaymentEntry.invoice_received_entry_id == invoice.id)
        .first()
    )
    if not payment:
        raise InvoiceFlowError("No linked Upcoming payment to return to.")

    files_moved: list[str] = []
    for att in list(invoice.attachments or []):
        s.add(
            PaymentEntryAttachment(
                payment_entry_id=payment.id,
                filename=att.filename,
                storage_key=att.storage_key,
                content_type=att.content_type,
                size_bytes=att.size_bytes,
                uploaded_by_user_id=att.uploaded_by_user_id,
            )
        )
        files_moved.append(att.filename)
        s.delete(att)
    invoice.attachments = []

    payment.invoice_received_entry_id = None
    meta = {
        "payment_entry_id": payment.id,
        "invoice_received_entry_id": invoice.id,
        "invoice": _invoice_row_snapshot(invoice),
        "payment": _payment_row_snapshot(payment),
        "files_moved": files_moved,
    }
    s.delete(invoice)
    s.flush()

    record_event(
        s,
        actor=user,
        action="payment_entry.returned_to_upcoming",
        entity_type="PaymentEntry",
        entity_id=str(payment.id),
        metadata=meta,
    )
    return payment


def match_invoice_to_po(s, *, invoice, purchase_order, user) -> None:
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry

    if invoice.disposition == InvoiceReceivedEntry.DISPOSITION_OTHER_PAYMENT:
        raise InvoiceFlowError("Unmark Other Payment before matching a PO.")

    before_po = invoice.purchase_order_id
    invoice.purchase_order_id = purchase_order.id
    invoice.disposition = InvoiceReceivedEntry.DISPOSITION_PO_MATCHED
    invoice.updated_at = utcnow()
    s.flush()
    record_event(
        s,
        actor=user,
        action="invoice_received.po_matched",
        entity_type="InvoiceReceivedEntry",
        entity_id=str(invoice.id),
        metadata={
            "invoice": _invoice_row_snapshot(invoice),
            "purchase_order_id": purchase_order.id,
            "po_number": purchase_order.po_number,
            "before_purchase_order_id": before_po,
            "files": _file_snapshot(invoice.attachments),
        },
    )


def unmatch_invoice_from_po(s, *, invoice, user) -> None:
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry

    before_po = invoice.purchase_order_id
    before_po_number = invoice.purchase_order.po_number if invoice.purchase_order else None
    invoice.purchase_order_id = None
    invoice.disposition = InvoiceReceivedEntry.DISPOSITION_UNASSIGNED
    invoice.updated_at = utcnow()
    s.flush()
    record_event(
        s,
        actor=user,
        action="invoice_received.po_unmatched",
        entity_type="InvoiceReceivedEntry",
        entity_id=str(invoice.id),
        metadata={
            "invoice": _invoice_row_snapshot(invoice),
            "purchase_order_id": before_po,
            "po_number": before_po_number,
            "files": _file_snapshot(invoice.attachments),
        },
    )


def mark_invoice_other_payment(s, *, invoice, user) -> None:
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry

    if invoice.purchase_order_id or invoice.disposition == InvoiceReceivedEntry.DISPOSITION_PO_MATCHED:
        raise InvoiceFlowError("Unmatch the purchase order before marking as Other Payment.")

    before = _invoice_row_snapshot(invoice)
    invoice.purchase_order_id = None
    invoice.disposition = InvoiceReceivedEntry.DISPOSITION_OTHER_PAYMENT
    invoice.updated_at = utcnow()
    s.flush()
    record_event(
        s,
        actor=user,
        action="invoice_received.marked_other",
        entity_type="InvoiceReceivedEntry",
        entity_id=str(invoice.id),
        metadata={
            "before": before,
            "after": _invoice_row_snapshot(invoice),
            "files": _file_snapshot(invoice.attachments),
        },
    )


def mark_invoice_paid(s, *, invoice, user) -> None:
    """Mark invoice paid (D55/D56/D57). Never closes a purchase order.

    - With a matched PO: set is_paid; stays po_matched for Related invoices.
    - With no PO: route to Other Payments via mark_invoice_other_payment, then set is_paid.
    """
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry

    before = _invoice_row_snapshot(invoice)
    if invoice.purchase_order_id or invoice.disposition == InvoiceReceivedEntry.DISPOSITION_PO_MATCHED:
        if not invoice.purchase_order_id:
            raise InvoiceFlowError("po_matched invoice is missing its purchase order link.")
        invoice.is_paid = True
        invoice.updated_at = utcnow()
        s.flush()
    else:
        # Explicit path to Other Payments (reuse existing helper), then flag paid.
        mark_invoice_other_payment(s, invoice=invoice, user=user)
        invoice.is_paid = True
        invoice.updated_at = utcnow()
        s.flush()

    record_event(
        s,
        actor=user,
        action="invoice_received.marked_paid",
        entity_type="InvoiceReceivedEntry",
        entity_id=str(invoice.id),
        metadata={
            "before": before,
            "after": _invoice_row_snapshot(invoice),
            "files": _file_snapshot(invoice.attachments),
            "purchase_order_id": invoice.purchase_order_id,
            "po_untouched": True,
        },
    )


def build_weekly_brief_payment_rows(s) -> list[dict]:
    """Outstanding obligations for the weekly brief: one row per payment/invoice.

    Migrated PaymentEntry rows (linked invoice) are excluded; paid invoices are excluded.
    Other-payment invoices still appear until marked paid (same as pre-P4-08A brief scope).
    """
    from sqlalchemy import nulls_last

    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry, PaymentEntry

    payments = (
        s.query(PaymentEntry)
        .filter(PaymentEntry.invoice_received_entry_id.is_(None))
        .order_by(nulls_last(PaymentEntry.payment_due_date.asc()))
        .all()
    )
    invoices_received = (
        s.query(InvoiceReceivedEntry)
        .filter(InvoiceReceivedEntry.is_paid.is_(False))
        .all()
    )
    payment_rows: list[dict] = []
    for e in payments:
        payment_rows.append({"sort_date": e.order_date, "kind": "upcoming", "entry": e})
    for e in invoices_received:
        payment_rows.append({"sort_date": e.date_received, "kind": "received", "entry": e})
    payment_rows.sort(
        key=lambda r: (r["sort_date"] is None, r["sort_date"] or date.max, r["kind"])
    )
    return payment_rows


def restore_payment_entry_from_audit_snapshot(
    s,
    *,
    snapshot: dict,
    source_event_id: int,
    user=None,
) -> "PaymentEntry":
    """Recreate a PaymentEntry from an audit delete snapshot (P4-08A / D54)."""
    from decimal import Decimal, InvalidOperation

    from app.eqms.modules.purchasing.models import PaymentEntry

    amount_raw = snapshot.get("amount")
    amount = None
    if amount_raw is not None and str(amount_raw).strip() != "":
        try:
            amount = Decimal(str(amount_raw))
        except (InvalidOperation, ValueError) as e:
            raise InvoiceFlowError(f"Invalid amount in audit snapshot: {amount_raw}") from e

    due_raw = snapshot.get("payment_due_date")
    due = None
    if due_raw:
        due = date.fromisoformat(str(due_raw)[:10])

    entry = PaymentEntry(
        vendor=(snapshot.get("vendor") or "").strip() or None,
        description=(snapshot.get("description") or "").strip() or None,
        amount=amount,
        payment_due_date=due,
        invoice_received_entry_id=None,
    )
    s.add(entry)
    s.flush()
    record_event(
        s,
        actor=user,
        action="payment_entry.restored",
        entity_type="PaymentEntry",
        entity_id=str(entry.id),
        metadata={
            "source_audit_event_id": source_event_id,
            "snapshot": {
                "vendor": entry.vendor,
                "description": entry.description,
                "amount": str(entry.amount) if entry.amount is not None else None,
                "payment_due_date": entry.payment_due_date.isoformat() if entry.payment_due_date else None,
            },
            "new_id": entry.id,
            "original_id": snapshot.get("id"),
        },
    )
    return entry


def return_invoice_to_received(s, *, invoice, user) -> None:
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry

    invoice.purchase_order_id = None
    invoice.disposition = InvoiceReceivedEntry.DISPOSITION_UNASSIGNED
    invoice.updated_at = utcnow()
    s.flush()
    record_event(
        s,
        actor=user,
        action="invoice_received.returned_to_received",
        entity_type="InvoiceReceivedEntry",
        entity_id=str(invoice.id),
        metadata={
            "invoice": _invoice_row_snapshot(invoice),
            "files": _file_snapshot(invoice.attachments),
        },
    )


def enforce_disposition_invariant(invoice) -> None:
    """po_matched requires a PO; other_payment requires none."""
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry

    if invoice.disposition == InvoiceReceivedEntry.DISPOSITION_PO_MATCHED and not invoice.purchase_order_id:
        raise InvoiceFlowError("po_matched requires a purchase order.")
    if invoice.disposition == InvoiceReceivedEntry.DISPOSITION_OTHER_PAYMENT and invoice.purchase_order_id:
        raise InvoiceFlowError("other_payment cannot have a purchase order.")
