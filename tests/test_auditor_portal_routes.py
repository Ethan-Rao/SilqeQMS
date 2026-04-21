from __future__ import annotations

import os
import sys
from pathlib import Path

import pytest
from werkzeug.security import generate_password_hash

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)


MIN_PDF = b"%PDF-1.4\n1 0 obj<<>>endobj\ntrailer<<>>\n%%EOF\n"


def _write_docx(path: Path, text: str) -> None:
    from docx import Document

    d = Document()
    d.add_paragraph(text)
    d.save(str(path))


def _write_xlsx(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "ColA"
    ws["B1"] = "ColB"
    ws["A2"] = "r1"
    ws["B2"] = "r2"
    wb.save(str(path))


@pytest.fixture()
def auditor_app_client(tmp_path, monkeypatch):
    from app.eqms import create_app
    from app.eqms.models import Permission, Role, User
    from app.eqms.modules.auditor_portal import fs as auditor_fs

    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("SECRET_KEY", "test-secret")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 't.db'}")
    monkeypatch.setenv("ENV", "test")
    monkeypatch.setenv("STORAGE_BACKEND", "local")
    monkeypatch.setenv("STORAGE_LOCAL_ROOT", str(tmp_path / "blobs"))
    for k in ("S3_ENDPOINT", "S3_REGION", "S3_BUCKET", "S3_ACCESS_KEY_ID", "S3_SECRET_ACCESS_KEY"):
        monkeypatch.delenv(k, raising=False)

    auditor_fs.reset_auditor_root_cache_for_tests()
    files_root = tmp_path / "Auditor Files"
    policies = files_root / "Policies"
    records = files_root / "Records"
    policies.mkdir(parents=True)
    records.mkdir(parents=True)
    _write_docx(policies / "policy.docx", "AuditorPolicyBody123")
    (policies / "old_spec.doc").write_bytes(b"0\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1\x00")  # minimal OLE header-ish
    _write_xlsx(records / "log.xlsx")
    (records / "report.pdf").write_bytes(MIN_PDF)

    app = create_app()
    app.config["_schema_health_ok"] = True
    app.config["AUDITOR_PORTAL_ENABLED"] = "1"
    app.config["AUDITOR_FILES_ROOT"] = str(files_root)
    app.config["AUDITOR_PDF_BACKEND"] = "xhtml2pdf"

    engine = app.extensions["sqlalchemy_engine"]
    from tests._auditor_schema import create_auditor_test_schema

    create_auditor_test_schema(engine, include_admin_docs=True)

    from app.eqms.db import session_scope

    with session_scope(app) as s:
        p_access = Permission(key="auditor_portal.access", name="Auditor Portal: access")
        p_admin_log = Permission(key="auditor_portal.admin", name="Auditor Portal: admin log")
        p_admin_view = Permission(key="admin.view", name="Admin: view")
        p_docs = Permission(key="docs.view", name="Docs")
        s.add_all([p_access, p_admin_log, p_admin_view, p_docs])
        s.flush()
        r_aud = Role(key="auditor", name="Auditor")
        r_aud.permissions.append(p_access)
        r_admin = Role(key="admin", name="Admin")
        r_admin.permissions.extend([p_admin_view, p_admin_log, p_docs])
        u_aud = User(email="auditor@test.com", password_hash=generate_password_hash("aud"), is_active=True)
        u_aud.roles.append(r_aud)
        u_adm = User(email="admin@test.com", password_hash=generate_password_hash("adm"), is_active=True)
        u_adm.roles.append(r_admin)
        s.add_all([r_aud, r_admin, u_aud, u_adm])

    client = app.test_client()
    yield client, app
    auditor_fs.reset_auditor_root_cache_for_tests()


def test_dashboard_and_browse(auditor_app_client):
    client, _app = auditor_app_client
    client.post("/auth/login", data={"email": "auditor@test.com", "password": "aud"})
    r = client.get("/auditor/")
    assert r.status_code == 200
    assert b"Policies" in r.data and b"Records" in r.data
    r2 = client.get("/auditor/browse/Policies")
    assert r2.status_code == 200
    assert b"policy.docx" in r2.data


def test_pdf_inline(auditor_app_client):
    client, _ = auditor_app_client
    client.post("/auth/login", data={"email": "auditor@test.com", "password": "aud"})
    r = client.get("/auditor/file/Records/report.pdf")
    assert r.status_code == 200
    assert r.headers.get("Content-Type", "").startswith("application/pdf")


def test_docx_html_and_pdf(auditor_app_client):
    client, _ = auditor_app_client
    client.post("/auth/login", data={"email": "auditor@test.com", "password": "aud"})
    r = client.get("/auditor/file/Policies/policy.docx")
    assert r.status_code == 200
    assert b"AuditorPolicyBody123" in r.data
    assert b"View PDF version" in r.data
    rp = client.get("/auditor/file/Policies/policy.docx?as=pdf")
    assert rp.status_code == 200
    assert rp.data[:4] == b"%PDF"


def test_xlsx_default_pdf_and_table(auditor_app_client):
    client, _ = auditor_app_client
    client.post("/auth/login", data={"email": "auditor@test.com", "password": "aud"})
    r = client.get("/auditor/file/Records/log.xlsx")
    assert r.status_code == 200
    assert r.headers.get("Content-Type", "").startswith("application/pdf")
    rt = client.get("/auditor/file/Records/log.xlsx?as=table")
    assert rt.status_code == 200
    assert b"ColA" in rt.data


def test_doc_not_viewable(auditor_app_client):
    client, _ = auditor_app_client
    client.post("/auth/login", data={"email": "auditor@test.com", "password": "aud"})
    r = client.get("/auditor/file/Policies/old_spec.doc")
    assert r.status_code == 200
    assert b".docx" in r.data


def test_no_perm_403(auditor_app_client):
    client, app = auditor_app_client
    from app.eqms.db import session_scope
    from app.eqms.models import Permission, Role, User

    with session_scope(app) as s:
        p_none = Permission(key="noop.test", name="noop")
        s.add(p_none)
        s.flush()
        r = Role(key="nobody", name="Nobody")
        r.permissions.append(p_none)
        u = User(email="nobody@test.com", password_hash=generate_password_hash("x"), is_active=True)
        u.roles.append(r)
        s.add_all([r, u])
    client.post("/auth/login", data={"email": "nobody@test.com", "password": "x"})
    assert client.get("/auditor/").status_code == 403


def test_flag_off_404(tmp_path, monkeypatch):
    from app.eqms import create_app
    from app.eqms.models import Permission, Role, User

    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("SECRET_KEY", "test-secret")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'a.db'}")
    monkeypatch.setenv("ENV", "test")
    monkeypatch.setenv("STORAGE_BACKEND", "local")
    monkeypatch.setenv("STORAGE_LOCAL_ROOT", str(tmp_path / "b"))
    app = create_app()
    app.config["_schema_health_ok"] = True
    app.config["AUDITOR_PORTAL_ENABLED"] = "0"
    from tests._auditor_schema import create_auditor_test_schema

    create_auditor_test_schema(app.extensions["sqlalchemy_engine"], include_admin_docs=False)
    from app.eqms.db import session_scope

    with session_scope(app) as s:
        pa = Permission(key="auditor_portal.access", name="a")
        ra = Role(key="auditor", name="Auditor")
        ra.permissions.append(pa)
        u = User(email="a@t.com", password_hash=generate_password_hash("p"), is_active=True)
        u.roles.append(ra)
        s.add_all([pa, ra, u])
    c = app.test_client()
    c.post("/auth/login", data={"email": "a@t.com", "password": "p"})
    assert c.get("/auditor/").status_code == 404


def test_admin_only_no_auditor_perm_403(auditor_app_client):
    client, _ = auditor_app_client
    client.post("/auth/login", data={"email": "admin@test.com", "password": "adm"})
    assert client.get("/auditor/").status_code == 403


def test_auditor_cannot_admin(auditor_app_client):
    client, _ = auditor_app_client
    client.post("/auth/login", data={"email": "auditor@test.com", "password": "aud"})
    assert client.get("/admin/").status_code == 403


def test_admin_still_reaches_admin_docs_view(auditor_app_client, tmp_path, monkeypatch):
    """Regression: admin shell document view unchanged."""
    from app.eqms.db import session_scope
    from app.eqms.modules.admin_docs.models import AdminDocFile, AdminDocFolder
    from app.eqms.storage import storage_from_config

    client, app = auditor_app_client
    store_root = tmp_path / "store"
    store_root.mkdir(parents=True)
    monkeypatch.setitem(app.config, "STORAGE_LOCAL_ROOT", str(store_root))

    with session_scope(app) as s:
        folder = AdminDocFolder(library_key="qms_documents", name="F", parent_id=None)
        s.add(folder)
        s.flush()
        doc = AdminDocFile(
            library_key="qms_documents",
            folder_id=folder.id,
            filename="inline.docx",
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            storage_key="admin_docs/inline.docx",
            size_bytes=10,
        )
        s.add(doc)
        s.flush()
        doc_id = doc.id

    p = tmp_path / "seed.docx"
    _write_docx(p, "AdminDocViewBody987")
    st = storage_from_config(app.config)
    st.put_bytes("admin_docs/inline.docx", p.read_bytes(), content_type=doc.content_type)

    client.post("/auth/login", data={"email": "admin@test.com", "password": "adm"})
    r = client.get(f"/admin/admin-docs/documents/{doc_id}/view")
    assert r.status_code == 200
    assert b"AdminDocViewBody987" in r.data
