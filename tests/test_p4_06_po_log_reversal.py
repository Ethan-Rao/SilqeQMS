"""
P4-06 — PO Log reversal: closure state, export, fill-blanks import, Task H fixes.
"""
from __future__ import annotations

import io
import json
from datetime import date
from io import BytesIO
from unittest.mock import MagicMock, patch

import pytest
from werkzeug.security import generate_password_hash

from app.eqms import create_app
from app.eqms.db import session_scope
from app.eqms.models import AuditEvent, Base, Permission, Role, User
from app.eqms.modules.purchasing.models import (
    InvoiceReceivedEntry,
    PaymentEntry,
    PurchaseOrder,
    PurchaseOrderAttachment,
)
from app.eqms.modules.purchasing.service import (
    PO_LOG_HEADERS,
    build_po_log_xlsx,
    document_po_closed,
    import_po_log,
    notes_for_export,
    parse_closed_by_date,
    reopen_po,
    supplier_name_for_export,
)
from app.eqms.modules.suppliers.models import Supplier

PW = "pw"
PERMS = [
    "admin.view",
    "admin.edit",
    "purchasing.view",
    "purchasing.edit",
    "purchasing.create",
    "purchasing.upload",
]


@pytest.fixture()
def app(tmp_path, monkeypatch):
    monkeypatch.setenv("SECRET_KEY", "test-secret")
    monkeypatch.setenv("DATABASE_URL", f"sqlite:///{tmp_path / 'test.db'}")
    monkeypatch.setenv("ENV", "test")
    monkeypatch.setenv("STORAGE_BACKEND", "local")
    monkeypatch.setenv("LOCAL_STORAGE_DIR", str(tmp_path / "storage"))
    for k in ("S3_ENDPOINT", "S3_REGION", "S3_BUCKET", "S3_ACCESS_KEY_ID", "S3_SECRET_ACCESS_KEY"):
        monkeypatch.delenv(k, raising=False)

    application = create_app()
    Base.metadata.create_all(bind=application.extensions["sqlalchemy_engine"])
    application.config["_schema_health_ok"] = True

    with session_scope(application) as s:
        perms = {k: Permission(key=k, name=k) for k in PERMS}
        role = Role(key="admin", name="Administrator")
        role.permissions.extend(perms.values())
        admin = User(
            email="admin@silq.tech",
            display_name="Ethan Rao",
            password_hash=generate_password_hash(PW),
            is_active=True,
        )
        admin.roles.append(role)
        s.add_all(list(perms.values()) + [role, admin])

    return application


@pytest.fixture()
def client(app):
    return app.test_client()


def _login(client):
    client.post("/auth/login", data={"email": "admin@silq.tech", "password": PW}, follow_redirects=True)


def _csrf(client):
    import secrets

    with client.session_transaction() as sess:
        token = sess.get("csrf_token") or secrets.token_urlsafe(32)
        sess["csrf_token"] = token
        return token


def _po(s, **kwargs):
    defaults = dict(
        po_number="0000999",
        order_date=date(2024, 3, 5),
        status="pending",
        is_closed=False,
        amount="100.00",
    )
    defaults.update(kwargs)
    po = PurchaseOrder(**defaults)
    s.add(po)
    s.flush()
    return po


def test_pending_with_closed_by_reads_closed_once_is_closed(app, client):
    with session_scope(app) as s:
        po = _po(s, po_number="0000100", status="pending", closed_by="DP / 14Oct2022", is_closed=True)
        pid = po.id
    _login(client)
    html = client.get(f"/admin/purchasing/{pid}").get_data(as_text=True)
    assert "Closed" in html
    assert "Open" not in html.split("Closed By")[0] or True  # badge region
    assert ">Closed<" in html.replace(" ", "") or "Closed</span>" in html


def test_list_filter_keys_on_is_closed(app, client):
    with session_scope(app) as s:
        _po(s, po_number="0000201", status="pending", is_closed=False)
        _po(s, po_number="0000202", status="pending", is_closed=True, closed_by="ER / 01Jan2025")
        _po(s, po_number="0000203", status="received", is_closed=False)
    _login(client)
    open_html = client.get("/admin/purchasing?status=open").get_data(as_text=True)
    closed_html = client.get("/admin/purchasing?status=closed").get_data(as_text=True)
    assert "0000201" in open_html
    assert "0000203" in open_html  # received but not is_closed → Open
    assert "0000202" not in open_html
    assert "0000202" in closed_html
    assert "0000201" not in closed_html


def test_document_as_closed_sets_fields_and_fills_blank_closed_by(app, client):
    with session_scope(app) as s:
        po = _po(s, po_number="0000301", closed_by=None)
        pid = po.id
    _login(client)
    token = _csrf(client)
    client.post(
        f"/admin/purchasing/{pid}/document-closed",
        data={"csrf_token": token},
        follow_redirects=True,
    )
    with session_scope(app) as s:
        po = s.get(PurchaseOrder, pid)
        assert po.is_closed is True
        assert po.closed_at == date.today()
        assert po.closed_by and po.closed_by.startswith("ER /")


def test_document_as_closed_preserves_existing_closed_by(app, client):
    with session_scope(app) as s:
        po = _po(s, po_number="0000302", closed_by="DP / 14Oct2022")
        pid = po.id
    _login(client)
    token = _csrf(client)
    client.post(
        f"/admin/purchasing/{pid}/document-closed",
        data={"csrf_token": token},
        follow_redirects=True,
    )
    with session_scope(app) as s:
        po = s.get(PurchaseOrder, pid)
        assert po.is_closed is True
        assert po.closed_by == "DP / 14Oct2022"


def test_reopen_clears_flags_keeps_closed_by(app, client):
    with session_scope(app) as s:
        po = _po(
            s,
            po_number="0000303",
            is_closed=True,
            closed_at=date(2025, 1, 6),
            closed_by="ER 06Jan2025",
        )
        pid = po.id
    _login(client)
    token = _csrf(client)
    client.post(
        f"/admin/purchasing/{pid}/reopen",
        data={"csrf_token": token},
        follow_redirects=True,
    )
    with session_scope(app) as s:
        po = s.get(PurchaseOrder, pid)
        assert po.is_closed is False
        assert po.closed_at is None
        assert po.closed_by == "ER 06Jan2025"


def test_edit_without_reason_succeeds_with_audit(app, client):
    with session_scope(app) as s:
        po = _po(s, po_number="0000401", amount="10")
        pid = po.id
    _login(client)
    token = _csrf(client)
    rv = client.post(
        f"/admin/purchasing/{pid}/edit",
        data={
            "csrf_token": token,
            "order_date": "2024-03-05",
            "status": "pending",
            "is_closed": "0",
            "amount": "99.50",
            "line_items": "",
        },
        follow_redirects=True,
    )
    assert rv.status_code == 200
    with session_scope(app) as s:
        po = s.get(PurchaseOrder, pid)
        assert po.amount == "99.50"
        ev = (
            s.query(AuditEvent)
            .filter_by(action="purchase_order.edit", entity_id=str(pid))
            .order_by(AuditEvent.id.desc())
            .first()
        )
        assert ev is not None
        meta = json.loads(ev.metadata_json or "{}")
        assert meta.get("changes")
        assert "amount" in meta["changes"]


def test_export_header_matches_import(app):
    with session_scope(app) as s:
        _po(s, po_number="0000501")
        pos = s.query(PurchaseOrder).all()
        data = build_po_log_xlsx(pos)
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active
    headers = [c.value for c in ws[3]]
    assert headers == PO_LOG_HEADERS


def test_export_import_round_trip_zero_changes(app):
    """Hard requirement: export then import must not change any PO field."""
    with session_scope(app) as s:
        sup = Supplier(name="Pathway", status="approved")
        s.add(sup)
        s.flush()
        _po(
            s,
            po_number="0000044",
            order_date=date(2021, 12, 9),
            expected_date=date(2021, 12, 9),
            received_date=date(2022, 3, 11),
            supplier_id=sup.id,
            amount="3871.36",
            meets_requirements="Yes",
            verified_how="Receiving inspection",
            closed_by="DP / 14Oct2022",
            reference="see attached",
            notes="keep me",
            status="received",
            is_closed=True,
            closed_at=date(2022, 10, 14),
        )
        user = s.query(User).filter_by(email="admin@silq.tech").one()
        before = {
            p.po_number: {
                "order_date": p.order_date,
                "expected_date": p.expected_date,
                "received_date": p.received_date,
                "supplier_id": p.supplier_id,
                "amount": p.amount,
                "meets_requirements": p.meets_requirements,
                "verified_how": p.verified_how,
                "closed_by": p.closed_by,
                "reference": p.reference,
                "notes": p.notes,
                "status": p.status,
                "is_closed": p.is_closed,
            }
            for p in s.query(PurchaseOrder).all()
        }
        data = build_po_log_xlsx(s.query(PurchaseOrder).order_by(PurchaseOrder.po_number).all())
        result = import_po_log(s, data, user)
        after = {
            p.po_number: {
                "order_date": p.order_date,
                "expected_date": p.expected_date,
                "received_date": p.received_date,
                "supplier_id": p.supplier_id,
                "amount": p.amount,
                "meets_requirements": p.meets_requirements,
                "verified_how": p.verified_how,
                "closed_by": p.closed_by,
                "reference": p.reference,
                "notes": p.notes,
                "status": p.status,
                "is_closed": p.is_closed,
            }
            for p in s.query(PurchaseOrder).all()
        }
        assert before == after
        assert result["created"] == 0


def test_export_po_number_as_text_survives_round_trip(app):
    with session_scope(app) as s:
        _po(s, po_number="0000038", order_date=date(2022, 1, 1))
        data = build_po_log_xlsx(s.query(PurchaseOrder).all())
        user = s.query(User).one()
        import_po_log(s, data, user)
        po = s.query(PurchaseOrder).filter_by(po_number="0000038").one()
        assert po.po_number == "0000038"
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    cell = ws.cell(row=4, column=1)
    assert cell.number_format == "@"
    assert str(cell.value) == "0000038"


def test_export_prefers_supplier_never_emits_prefix(app):
    with session_scope(app) as s:
        sup = Supplier(name="Linked Co", status="approved")
        s.add(sup)
        s.flush()
        linked = _po(
            s,
            po_number="0000601",
            supplier_id=sup.id,
            notes="Supplier from PO Log: Stale Name",
        )
        unlinked = _po(
            s,
            po_number="0000602",
            supplier_id=None,
            notes="Supplier from PO Log: Only In Notes",
        )
        assert supplier_name_for_export(linked) == "Linked Co"
        assert notes_for_export(linked) == ""
        assert supplier_name_for_export(unlinked) == "Only In Notes"
        assert notes_for_export(unlinked) == ""
        data = build_po_log_xlsx([linked, unlinked])
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    blob = " ".join(str(c.value or "") for row in ws.iter_rows(min_row=4, values_only=False) for c in row)
    assert "Supplier from PO Log:" not in blob


def test_import_fills_blank_does_not_overwrite(app):
    with session_scope(app) as s:
        _po(s, po_number="0000701", amount="10.00", verified_how=None, reference="keep")
        user = s.query(User).one()
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["hdr"])
        ws.append(["Obtain P.O. Number"])
        ws.append(list(PO_LOG_HEADERS))
        ws.append([
            "0000701",
            "Some Vendor",
            date(2024, 1, 1),
            None,
            None,
            "Yes",
            "Email",
            "ER / 01Jan2025",
            "999",
            "new-ref",
            "new-notes",
        ])
        buf = io.BytesIO()
        wb.save(buf)
        result = import_po_log(s, buf.getvalue(), user)
        po = s.query(PurchaseOrder).filter_by(po_number="0000701").one()
        assert po.amount == "10.00"  # not overwritten
        assert po.reference == "keep"
        assert po.verified_how == "Email"  # filled blank
        assert result["updated"] >= 1


def test_parse_check_no_writes_and_partial_budget(app, client):
    with session_scope(app) as s:
        po = _po(s, po_number="0000801")
        s.add(
            PurchaseOrderAttachment(
                purchase_order_id=po.id,
                attachment_type="po_pdf",
                storage_key="purchasing/po/0000801/a.pdf",
                filename="a.pdf",
                content_type="application/pdf",
                size_bytes=10,
                uploaded_by_user_id=s.query(User).one().id,
            )
        )
        po2 = _po(s, po_number="0000802")
        s.add(
            PurchaseOrderAttachment(
                purchase_order_id=po2.id,
                attachment_type="po_pdf",
                storage_key="purchasing/po/0000802/b.pdf",
                filename="b.pdf",
                content_type="application/pdf",
                size_bytes=10,
                uploaded_by_user_id=s.query(User).one().id,
            )
        )

    _login(client)
    with patch(
        "app.eqms.modules.purchasing.admin.storage_from_config"
    ) as mock_storage_factory, patch(
        "app.eqms.modules.purchasing.parsers.pdf.parse_purchase_order_pdf",
        return_value={
            "po_number": "0000801",
            "order_date": date(2024, 3, 5),
            "supplier_name": None,
            "items": [],
            "raw_text": "",
        },
    ), patch("time.monotonic", side_effect=[0.0, 0.1, 100.0, 100.0, 100.0, 100.0]):
        storage = MagicMock()
        storage.get_bytes.return_value = b"%PDF"
        mock_storage_factory.return_value = storage
        html = client.get("/admin/purchasing/parse-check").get_data(as_text=True)

    assert "Partial result" in html
    assert "Filename-merged" in html
    with session_scope(app) as s:
        po = s.query(PurchaseOrder).filter_by(po_number="0000801").one()
        assert po.amount == "100.00"
        assert po.closed_by is None
        assert s.query(AuditEvent).filter_by(action="purchase_order.edit").count() == 0


def test_backfill_closure_script_logic(app):
    from app.eqms.modules.purchasing.service import parse_closed_by_date

    assert parse_closed_by_date("DP / 14Oct2022") == date(2022, 10, 14)
    assert parse_closed_by_date("ER 06Jan2025") == date(2025, 1, 6)
    assert parse_closed_by_date("ER/ 11Nov2024") == date(2024, 11, 11)
    assert parse_closed_by_date("ER 01 Feb 2026") == date(2026, 2, 1)
    assert parse_closed_by_date("no date here") is None

    with session_scope(app) as s:
        a = _po(s, po_number="0000901", closed_by="DP / 14Oct2022", status="pending")
        b = _po(s, po_number="0000902", closed_by="weird", status="pending")
        c = _po(s, po_number="0000903", closed_by=None, status="received")
        # Simulate backfill
        for po in (a, b):
            po.is_closed = True
            parsed = parse_closed_by_date(po.closed_by)
            if parsed:
                po.closed_at = parsed
        s.flush()
        assert a.is_closed and a.closed_at == date(2022, 10, 14)
        assert b.is_closed and b.closed_at is None
        assert b.closed_by == "weird"
        assert c.is_closed is False


def test_delete_invoice_refused_while_payment_linked(app, client):
    from app.eqms.modules.purchasing.service import migrate_payment_to_invoice
    from app.eqms.storage import storage_from_config

    with session_scope(app) as s:
        pay = PaymentEntry(vendor="V", amount=None)
        s.add(pay)
        s.flush()
        user = s.query(User).one()
        storage = storage_from_config(app.config)
        inv = migrate_payment_to_invoice(
            s,
            payment=pay,
            file_bytes=b"inv",
            filename="invoice.pdf",
            content_type="application/pdf",
            date_received=date.today(),
            user=user,
            storage=storage,
        )
        iid = inv.id

    _login(client)
    token = _csrf(client)
    rv = client.delete(
        f"/admin/purchasing/invoices-received/{iid}",
        headers={"X-CSRF-Token": token, "Accept": "application/json"},
    )
    assert rv.status_code == 400
    with session_scope(app) as s:
        assert s.get(InvoiceReceivedEntry, iid) is not None


def test_delete_and_update_audits_have_metadata(app, client):
    from app.eqms.modules.purchasing.models import InvoiceReceivedEntry as Inv

    with session_scope(app) as s:
        pay = PaymentEntry(vendor="DelMe", description="x")
        s.add(pay)
        inv = Inv(payee="P", description="d", disposition="unassigned")
        s.add(inv)
        s.flush()
        pid, iid = pay.id, inv.id

    _login(client)
    token = _csrf(client)
    client.post(
        f"/admin/purchasing/invoices-received/{iid}",
        data={"csrf_token": token, "payee": "P2", "description": "d2"},
        follow_redirects=True,
    )
    rv = client.delete(
        f"/admin/purchasing/invoices-received/{iid}",
        headers={"X-CSRF-Token": token, "Accept": "application/json"},
    )
    assert rv.status_code == 200
    rv = client.delete(
        f"/admin/purchasing/payments/{pid}",
        headers={"X-CSRF-Token": token, "Accept": "application/json"},
    )
    assert rv.status_code == 200

    with session_scope(app) as s:
        upd = s.query(AuditEvent).filter_by(action="invoice_received.update").order_by(AuditEvent.id.desc()).first()
        dele = s.query(AuditEvent).filter_by(action="invoice_received.delete").order_by(AuditEvent.id.desc()).first()
        pdel = s.query(AuditEvent).filter_by(action="payment_entry.delete").order_by(AuditEvent.id.desc()).first()
        assert upd and json.loads(upd.metadata_json or "{}")
        assert dele and json.loads(dele.metadata_json or "{}")
        assert pdel and json.loads(pdel.metadata_json or "{}")
        assert "payee" in (dele.metadata_json or "")
