"""
Parse ``SILQ Training Matrix.xlsx`` into a portable JSON reference.

Phase 7 (Prompt 15 Task C1). READ-ONLY utility — no DB, no prod credentials.
Committed alongside its output (docs/training_matrix_parsed.json) so future
bulk-assign runs have a stable reference of "which document is required for which
role" without re-reading the spreadsheet.

The ``Matrix`` sheet layout:
- Row 1: role/employee labels in columns C onward (columns A/B are blank).
- Column B: the training item (e.g. "QM.SLQ001 Document Control").
- Body cells: "X"/"x" marks a required item; blank means not required.

Usage:
    python scripts/parse_training_matrix.py
    python scripts/parse_training_matrix.py --xlsx "path/to/SILQ Training Matrix.xlsx"
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "SILQ Training Matrix.xlsx"
OUTPUT = ROOT / "docs" / "training_matrix_parsed.json"

# Leading controlled-document code, e.g. "QM.SLQ001", "FM1-QM.SLQ016".
_DOC_RE = re.compile(r"^\s*([A-Za-z0-9]+(?:-[A-Za-z0-9.]+)?\.SLQ\d+)", re.IGNORECASE)


def _doc_number(item_text: str) -> str:
    """Extract a controlled-doc code from a matrix row label, else the raw text."""
    m = _DOC_RE.match(item_text or "")
    return m.group(1).upper() if m else (item_text or "").strip()


def parse_matrix(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Matrix" not in wb.sheetnames:
        raise SystemExit(f"No 'Matrix' sheet in {xlsx_path.name}; sheets: {wb.sheetnames}")
    ws = wb["Matrix"]

    # Employee/role columns: row 1, from column C (index 3) rightward.
    employees: dict[int, str] = {}
    for col in range(3, ws.max_column + 1):
        label = ws.cell(row=1, column=col).value
        if label and str(label).strip():
            employees[col] = str(label).strip()

    assignments = []
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=2).value
        if not item or not str(item).strip():
            continue  # section header / blank row
        item = str(item).strip()
        required_for = []
        for col, name in employees.items():
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip().lower() == "x":
                required_for.append(name)
        if not required_for:
            continue
        assignments.append(
            {"doc_number": _doc_number(item), "item": item, "required_for": required_for}
        )

    return {"employees": list(employees.values()), "assignments": assignments}


def main():
    ap = argparse.ArgumentParser(description="Parse the SILQ Training Matrix into JSON.")
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to SILQ Training Matrix.xlsx")
    ap.add_argument("--out", default=str(OUTPUT), help="Output JSON path")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Not found: {xlsx_path}")

    data = parse_matrix(xlsx_path)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Employees ({len(data['employees'])}): {', '.join(data['employees'])}")
    print(f"Assignments (rows with >=1 requirement): {len(data['assignments'])}")
    print(f"Wrote {out_path.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
